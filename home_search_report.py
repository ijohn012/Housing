#!/usr/bin/env python3
"""
Weekly home-search report.

Pulls active for-sale listings from RentCast for Severna Park, Arnold and
Ellicott City MD, scores them against hard criteria, estimates value, and
emails a datestamped Excel workbook of the ten best candidates.

Secrets come from environment variables (GitHub Actions encrypted secrets).
Nothing sensitive is ever written to the repo.
"""

import json
import math
import os
import smtplib
import sys
from datetime import date, datetime, timezone
from email.message import EmailMessage
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------- CONFIG

BASE_URL = "https://api.rentcast.io/v1"

SEARCH_AREAS = [
    {"city": "Severna Park", "state": "MD", "psf": 295, "avg_dom": 37},
    {"city": "Arnold",       "state": "MD", "psf": 269, "avg_dom": 25},
    {"city": "Ellicott City", "state": "MD", "psf": 265, "avg_dom": 22},
]

# Hard criteria. Listings that fail are still reported, with a per-item flag.
MAX_PRICE = 650_000
MIN_BEDS = 3
MIN_BATHS = 3.0
MIN_SQFT = 2_000
PREFERRED_DRIVE_MIN = 30   # over this = FAR (ranks lower, still qualifies)
MAX_DRIVE_MIN = 40         # over this = hard fail
PROPERTY_TYPES = {"Single Family"}

# Commute anchor. Approximate coordinates for 7765 Freetown Rd, Glen Burnie MD.
# Replace with exact coordinates from Google Maps if you want tighter numbers.
OFFICE_LAT = 39.1662
OFFICE_LON = -76.6055
# CALIBRATION TABLE -- the important part.
# A haversine proxy cannot know that MD-100 runs straight west to Ellicott City
# while getting to Arnold means working around the Severn. So don't make it
# guess: look each of these up ONCE in Google Maps (typical weekday morning,
# from 7765 Freetown Rd to the town center listed) and put the real number in.
# The proxy is then only used to adjust for how far a specific listing sits
# from that town center, which is a job it can actually do.
#
#   city -> (real drive minutes, town-center lat, town-center lon)
CITY_DRIVE_CALIBRATION = {
    "Severna Park":  (16, 39.0704, -76.5455),   # <- REPLACE with your Maps number
    "Arnold":        (22, 39.0334, -76.5033),   # <- REPLACE with your Maps number
    "Ellicott City": (28, 39.2673, -76.7983),   # <- REPLACE with your Maps number
}

ROAD_FACTOR = 1.35       # straight-line miles -> road miles
LOCAL_MILES = 5.0        # first few miles are surface streets
LOCAL_MPH = 30.0
HIGHWAY_MPH = 50.0       # I-97 / MD-10 / MD-100 / US-50
SPREAD_MPH = 40.0        # speed used for within-city offsets from town center

# Quota control. Free tier is 50 requests/month per key.
# Sized for ONE key on the free plan, worst case (a 5-Saturday month):
#   3 listing calls x 5 runs        = 15
#   4 AVM value calls x 5 runs      = 20
#   3 property-record calls, monthly = 3
#                                    ---
#                                     38, leaving room for manual test runs.
MONTHLY_QUOTA_PER_KEY = 48   # hard stop 2 under the plan limit
MAX_AVM_CALLS_PER_RUN = 4

# Optional subdivision filters. Leave empty to accept everything.
# Municipal boundaries are more reliable than Zillow/Redfin neighborhood
# polygons, but subdivision is the field to trust when you know the area.
SUBDIVISION_ALLOWLIST: set[str] = set()
SUBDIVISION_BLOCKLIST: set[str] = set()

TOP_N = 10
REPORTS_DIR = Path("reports")
QUOTA_FILE = REPORTS_DIR / "quota.json"

# ------------------------------------------------------------ QUOTA STATE


def load_quota() -> dict:
    month = date.today().strftime("%Y-%m")
    if QUOTA_FILE.exists():
        data = json.loads(QUOTA_FILE.read_text())
        if data.get("month") == month:
            return data
    return {"month": month, "key1": 0, "key2": 0, "comps_pulled": False}


def save_quota(state: dict) -> None:
    REPORTS_DIR.mkdir(exist_ok=True)
    QUOTA_FILE.write_text(json.dumps(state, indent=2))


class RentCast:
    """Thin client with two-key failover and a hard monthly request cap."""

    def __init__(self, keys: list[str], quota: dict):
        self.keys = [k for k in keys if k]
        self.quota = quota
        self.run_calls = 0
        if not self.keys:
            raise SystemExit("No RentCast API key found. Set RENTCAST_API_KEY.")

    def _slot(self, i: int) -> str:
        return "key1" if i == 0 else "key2"

    def _available(self, i: int) -> bool:
        return self.quota[self._slot(i)] < MONTHLY_QUOTA_PER_KEY

    def get(self, path: str, params: dict) -> list | dict | None:
        last_error = None
        for i, key in enumerate(self.keys):
            if not self._available(i):
                print(f"  key {i+1}: monthly cap reached, skipping")
                continue
            try:
                r = requests.get(
                    f"{BASE_URL}{path}",
                    params=params,
                    headers={"Accept": "application/json", "X-Api-Key": key},
                    timeout=30,
                )
                self.quota[self._slot(i)] += 1
                self.run_calls += 1
                if r.status_code == 200:
                    return r.json()
                if r.status_code in (401, 402, 429):
                    print(f"  key {i+1}: HTTP {r.status_code}, failing over")
                    last_error = r.status_code
                    continue
                print(f"  {path}: HTTP {r.status_code} — {r.text[:180]}")
                return None
            except requests.RequestException as exc:
                last_error = exc
                print(f"  key {i+1}: {exc}")
                continue
        print(f"  all keys exhausted or failing ({last_error})")
        return None


# ------------------------------------------------------------- SCORING


def haversine(lat1, lon1, lat2, lon2) -> float:
    r = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def raw_proxy_minutes(lat, lon) -> float:
    """Fallback only: distance-based estimate for cities with no calibration."""
    miles = haversine(OFFICE_LAT, OFFICE_LON, lat, lon) * ROAD_FACTOR
    local = min(miles, LOCAL_MILES)
    highway = max(0.0, miles - LOCAL_MILES)
    return (local / LOCAL_MPH + highway / HIGHWAY_MPH) * 60


def drive_minutes(lat, lon, city: str | None = None) -> float | None:
    """Calibrated drive time: a real Maps number per city, adjusted for how far
    this listing sits from that city's center."""
    if lat is None or lon is None:
        return None
    cal = CITY_DRIVE_CALIBRATION.get(city or "")
    if not cal:
        return round(raw_proxy_minutes(lat, lon), 1)
    base_min, c_lat, c_lon = cal
    # How much farther (or nearer) the office this listing is than the center.
    d_listing = haversine(OFFICE_LAT, OFFICE_LON, lat, lon)
    d_center = haversine(OFFICE_LAT, OFFICE_LON, c_lat, c_lon)
    offset_miles = (d_listing - d_center) * ROAD_FACTOR
    return round(max(1.0, base_min + offset_miles / SPREAD_MPH * 60), 1)


def evaluate(listing: dict, area: dict) -> dict:
    beds = listing.get("bedrooms")
    baths = listing.get("bathrooms")
    sqft = listing.get("squareFootage")
    price = listing.get("price")
    ptype = listing.get("propertyType") or ""
    subdivision = (listing.get("subdivision") or "").strip()
    drive = drive_minutes(listing.get("latitude"), listing.get("longitude"),
                          listing.get("city"))

    dom = listing.get("daysOnMarket")
    if dom is None and listing.get("listedDate"):
        try:
            listed = datetime.fromisoformat(
                listing["listedDate"].replace("Z", "+00:00"))
            dom = (datetime.now(timezone.utc) - listed).days
        except ValueError:
            dom = None

    est_value = sqft * area["psf"] if sqft else None
    discount_pct = None
    if est_value and price:
        discount_pct = (est_value - price) / est_value

    flags = {
        "Beds OK": "PASS" if (beds or 0) >= MIN_BEDS else "FAIL",
        "Baths OK": ("PASS" if (baths or 0) >= MIN_BATHS
                     else "CLOSE" if (baths or 0) >= MIN_BATHS - 0.5 else "FAIL"),
        "SqFt OK": "PASS" if (sqft or 0) >= MIN_SQFT else "FAIL",
        "Price OK": "PASS" if price and price <= MAX_PRICE else "FAIL",
        "Drive Tier": ("FAIL" if drive is None or drive > MAX_DRIVE_MIN
                       else "PASS" if drive <= PREFERRED_DRIVE_MIN else "FAR"),
        "Type OK": "PASS" if ptype in PROPERTY_TYPES else "FAIL",
    }
    if subdivision:
        if SUBDIVISION_ALLOWLIST and subdivision not in SUBDIVISION_ALLOWLIST:
            flags["Subdivision OK"] = "FAIL"
        elif subdivision in SUBDIVISION_BLOCKLIST:
            flags["Subdivision OK"] = "FAIL"
        else:
            flags["Subdivision OK"] = "PASS"
    else:
        flags["Subdivision OK"] = "UNKNOWN"

    hard = [flags[k] for k in ("Beds OK", "SqFt OK", "Price OK", "Type OK")]
    hard.append("PASS" if flags["Drive Tier"] != "FAIL" else "FAIL")
    if all(f == "PASS" for f in hard) and flags["Baths OK"] == "PASS":
        verdict = "MEETS ALL"
    elif all(f == "PASS" for f in hard) and flags["Baths OK"] == "CLOSE":
        verdict = "NEAR MISS"
    else:
        verdict = "NO"

    return {
        "Address": listing.get("formattedAddress") or listing.get("addressLine1"),
        "City": listing.get("city"),
        "Subdivision": subdivision or "—",
        "Price": price,
        "Beds": beds,
        "Baths": baths,
        "SqFt": sqft,
        "Lot (acres)": round(listing["lotSize"] / 43560, 2) if listing.get("lotSize") else None,
        "Year Built": listing.get("yearBuilt"),
        "DOM": dom,
        "Drive (min est)": drive,
        "Est. Value": round(est_value) if est_value else None,
        "Discount %": round(discount_pct, 4) if discount_pct is not None else None,
        "AVM Value": None,
        "AVM Discount %": None,
        "Meets Criteria": verdict,
        **flags,
        "Area Avg DOM": area["avg_dom"],
        "Listing URL": f"https://www.zillow.com/homes/{(listing.get('formattedAddress') or '').replace(' ', '-')}_rb/",
    }


def sort_key(row: dict):
    rank = {"MEETS ALL": 0, "NEAR MISS": 1, "NO": 2}[row["Meets Criteria"]]
    tier = {"PASS": 0, "FAR": 1, "FAIL": 2}[row["Drive Tier"]]
    disc = row["AVM Discount %"] if row["AVM Discount %"] is not None else row["Discount %"]
    return (rank, tier, -(disc if disc is not None else -99), -(row["DOM"] or 0))


# ------------------------------------------------------------- WORKBOOK

HDRFILL = PatternFill("solid", fgColor="1F3864")
GOODFILL = PatternFill("solid", fgColor="C6EFCE")
NEARFILL = PatternFill("solid", fgColor="FFEB9C")
BADFILL = PatternFill("solid", fgColor="F2F2F2")


def write_sheet(ws, rows: list[dict], title: str):
    ws.cell(row=1, column=1, value=title).font = Font(name="Arial", size=13, bold=True)
    if not rows:
        ws.cell(row=3, column=1, value="No listings returned for this run.").font = Font(
            name="Arial", size=10, italic=True)
        return
    headers = list(rows[0].keys())
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = HDRFILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(1 + i)].width = max(11, min(38, len(h) + 6))
    ws.row_dimensions[3].height = 32
    for r, row in enumerate(rows, start=4):
        fill = {"MEETS ALL": GOODFILL, "NEAR MISS": NEARFILL}.get(
            row.get("Meets Criteria"), BADFILL)
        for i, h in enumerate(headers):
            c = ws.cell(row=r, column=1 + i, value=row[h])
            c.font = Font(name="Arial", size=10)
            if h in ("Price", "Est. Value", "AVM Value"):
                c.number_format = '$#,##0'
            if h in ("Discount %", "AVM Discount %"):
                c.number_format = '0.0%'
            if h == "Meets Criteria":
                c.fill = fill
                c.font = Font(name="Arial", size=10, bold=True)
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(rows) + 3}"


def build_report(top: list[dict], everything: list[dict], comps: list[dict] | None,
                 stamp: str, calls: int) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 10"
    write_sheet(ws, top, f"Top {len(top)} candidates — {stamp}")

    ws2 = wb.create_sheet("All Listings")
    write_sheet(ws2, everything, f"Everything returned — {stamp}")

    if comps:
        ws3 = wb.create_sheet("New Comps")
        write_sheet(ws3, comps, f"Recent sales — paste into your workbook's Comps tab")

    ws4 = wb.create_sheet("Run Notes")
    notes = [
        f"Run date: {stamp}",
        f"RentCast API requests used this run: {calls}",
        "",
        "Est. Value is squarefootage x the seeded median $/sqft for that city.",
        "It assumes average condition and no renovation. It is a screen, not an",
        "appraisal — the real number comes from your workbook once you rate the",
        "photos.",
        "",
        "Drive (min est) starts from a real Google Maps time to each town",
        "center (the CITY_DRIVE_CALIBRATION table in the script) and adjusts for",
        "how far this specific listing sits from that center. It is only as good",
        "as the numbers you put in that table -- check them once, then trust it.",
        "",
        "Subdivision UNKNOWN means RentCast returned no subdivision field for",
        "that record. Verify it from the listing itself before trusting the city",
        "label — map boundaries misplace houses routinely.",
    ]
    for i, n in enumerate(notes):
        ws4.cell(row=1 + i, column=1, value=n).font = Font(name="Arial", size=10)
    ws4.column_dimensions['A'].width = 80

    REPORTS_DIR.mkdir(exist_ok=True)
    path = REPORTS_DIR / f"home_search_{stamp}.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------- EMAIL


def summarize(top: list[dict], stamp: str, calls: int) -> tuple[str, str]:
    """Returns (subject, markdown body) used by both email and the GitHub issue."""
    meets = [r for r in top if r["Meets Criteria"] == "MEETS ALL"]
    subject = (f"Home search — {len(meets)} match"
               f"{'' if len(meets) == 1 else 'es'} — {stamp}")

    lines = [f"**{len(meets)} listing(s) meet every hard criterion.**", ""]
    if top:
        lines += ["| | Price | Beds/Baths | SqFt | DOM | Drive | vs Est. | Address |",
                  "|---|---|---|---|---|---|---|---|"]
        for r in top[:8]:
            disc = r["AVM Discount %"] if r["AVM Discount %"] is not None else r["Discount %"]
            lines.append(
                f"| {r['Meets Criteria']} | ${(r['Price'] or 0):,} | "
                f"{r['Beds']}/{r['Baths']} | {r['SqFt']} | {r['DOM']} | "
                f"~{r['Drive (min est)']}m | "
                f"{'—' if disc is None else f'{disc:+.1%}'} | "
                f"{r['Address']} ({r['Subdivision']}) |")
    else:
        lines.append("_No listings returned this run._")
    lines += ["", f"API requests used this run: **{calls}**", "",
              "Full workbook is attached to this run as an artifact, and committed "
              "under `/reports`."]
    return subject, "\n".join(lines)


def write_summary(subject: str, body: str) -> None:
    REPORTS_DIR.mkdir(exist_ok=True)
    (REPORTS_DIR / "latest_subject.txt").write_text(subject)
    (REPORTS_DIR / "latest_summary.md").write_text(body)


def send_email(path: Path, subject: str, body: str) -> None:
    addr = os.environ.get("EMAIL_ADDRESS")
    pw = os.environ.get("EMAIL_APP_PASSWORD")
    if not addr or not pw:
        print("No Gmail app password set — skipping email. "
              "The report is committed to /reports, uploaded as a run artifact, "
              "and summarized in a GitHub issue.")
        return

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = addr
    msg["To"] = addr
    msg.set_content(body)
    msg.add_attachment(
        path.read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(addr, pw)
        s.send_message(msg)
    print(f"Emailed {path.name} to {addr}")


# ----------------------------------------------------------------- MAIN


def main() -> int:
    # Biweekly toggle: cron cannot express "every other week", so skip odd weeks.
    if os.environ.get("RUN_BIWEEKLY", "").lower() in ("1", "true", "yes"):
        week = date.today().isocalendar().week
        if week % 2 == 1:
            print(f"Biweekly mode, ISO week {week} is odd — skipping this run.")
            return 0

    quota = load_quota()
    api = RentCast([os.environ.get("RENTCAST_API_KEY", ""),
                    os.environ.get("RENTCAST_API_KEY_2", "")], quota)

    stamp = date.today().isoformat()
    everything: list[dict] = []

    for area in SEARCH_AREAS:
        print(f"Fetching {area['city']}...")
        data = api.get("/listings/sale", {
            "city": area["city"],
            "state": area["state"],
            "status": "Active",
            "propertyType": "Single Family",
            "limit": 200,
        })
        if not data:
            continue
        listings = data if isinstance(data, list) else data.get("listings", [])
        print(f"  {len(listings)} raw listings")
        for lst in listings:
            everything.append(evaluate(lst, area))

    everything.sort(key=sort_key)

    # Value estimates, capped, spent only on the strongest candidates.
    budget = min(MAX_AVM_CALLS_PER_RUN, len(everything))
    spent = 0
    for row in everything:
        if spent >= budget or row["Meets Criteria"] == "NO":
            break
        avm = api.get("/avm/value", {
            "address": row["Address"],
            "propertyType": "Single Family",
            "bedrooms": row["Beds"],
            "bathrooms": row["Baths"],
            "squareFootage": row["SqFt"],
        })
        spent += 1
        if avm and avm.get("price"):
            row["AVM Value"] = round(avm["price"])
            if row["Price"]:
                row["AVM Discount %"] = round(
                    (avm["price"] - row["Price"]) / avm["price"], 4)
    print(f"AVM calls spent: {spent}")

    everything.sort(key=sort_key)
    top = everything[:TOP_N]

    # First run of the month: pull recent sales for the Comps tab.
    comps = None
    if not quota.get("comps_pulled"):
        comps = []
        for area in SEARCH_AREAS:
            data = api.get("/properties", {
                "city": area["city"],
                "state": area["state"],
                "propertyType": "Single Family",
                "saleDateRange": 180,
                "bedrooms": 3,
                "limit": 40,
            })
            if not data:
                continue
            records = data if isinstance(data, list) else data.get("properties", [])
            for p in records:
                sqft = p.get("squareFootage")
                sold = p.get("lastSalePrice")
                if not sqft or not sold:
                    continue
                comps.append({
                    "Address": p.get("formattedAddress"),
                    "Area": area["city"],
                    "Subdivision": p.get("subdivision") or "—",
                    "Sold Price": sold,
                    "SqFt": sqft,
                    "$/sqft": round(sold / sqft),
                    "Sold Date": (p.get("lastSaleDate") or "")[:10],
                    "Beds": p.get("bedrooms"),
                    "Baths": p.get("bathrooms"),
                    "Year Built": p.get("yearBuilt"),
                })
        comps.sort(key=lambda c: c["Sold Date"], reverse=True)
        quota["comps_pulled"] = True
        print(f"Monthly comps pull: {len(comps)} sales")

    path = build_report(top, everything, comps, stamp, api.run_calls)
    print(f"Wrote {path}")
    print(f"REQUESTS THIS RUN: {api.run_calls}  |  "
          f"month to date — key1 {quota['key1']}, key2 {quota['key2']}")

    subject, body = summarize(top, stamp, api.run_calls)
    write_summary(subject, body)

    save_quota(quota)
    send_email(path, subject, body)
    return 0


if __name__ == "__main__":
    sys.exit(main())
